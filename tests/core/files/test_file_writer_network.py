"""
Интеграционные тесты FileWriter: реальная загрузка изображений по публичным URL.

Проверяется каждый поддерживаемый целевой формат из markdown-пайплайна.
Для .md и .txt контент сохраняется как текст (URL в разметке остаётся в файле, без HTTP).

Запуск только этих тестов:
  uv run pytest tests/core/files/test_file_writer_network.py -v

Исключить из прогона (например в изолированном CI без интернета):
  uv run pytest tests/ -m "not network"

Артефакты: каталог tests/core/files/file_writer_output/ очищается в начале сессии
  (conftest), затем файлы записываются заново.
"""

from __future__ import annotations

import re
import zipfile
from io import BytesIO

import httpx
import pytest
from openpyxl import load_workbook

from core.files.writer import FileWriter
from core.files.writer.models import WriteOptions
from tests.core.files.file_writer_artifacts import overwrite_artifact

# Стабильные публичные JPEG по фиксированным id (редирект → финальный URL, follow_redirects).
IMG_HTTP_JPEG_1 = "https://picsum.photos/id/237/200/300"
IMG_HTTP_JPEG_2 = "https://picsum.photos/id/1084/200/300"

_NETWORK_OPTS = WriteOptions(
    http_timeout_seconds=90.0,
    max_image_bytes=30 * 1024 * 1024,
)


def _looks_like_image_body(data: bytes) -> bool:
    if len(data) < 12:
        return False
    if len(data) >= 8 and data[:8] == b"\x89PNG\r\n\x1a\n":
        return True
    if data[:4] == b"\x89PNG":
        return True
    if data[:3] == b"\xff\xd8\xff":
        return True
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return True
    if data[:4] == b"RIFF" and len(data) >= 12 and data[8:12] == b"WEBP":
        return True
    if data[:2] == b"BM":
        return True
    return False


def _count_data_uri_images(html: str) -> int:
    return len(re.findall(r"data:image/[^;]+;base64,", html, flags=re.IGNORECASE))


@pytest.fixture(scope="module")
def public_http_image_urls_reachable() -> None:
    """Проверка доступности обоих URL до тяжёлых конвертаций."""
    with httpx.Client(timeout=90.0, follow_redirects=True) as client:
        for url in (IMG_HTTP_JPEG_1, IMG_HTTP_JPEG_2):
            response = client.get(url)
            response.raise_for_status()
            body = response.content
            assert len(body) > 200, f"{url}: слишком короткий ответ"
            assert _looks_like_image_body(body), (
                f"{url}: ответ не похож на изображение, начало {body[:24]!r}"
            )


@pytest.fixture
def network_writer() -> FileWriter:
    return FileWriter(options=_NETWORK_OPTS)


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_probe_public_http_image_urls(public_http_image_urls_reachable) -> None:
    """Явный тест: оба URL отдают изображение по HTTP."""
    assert public_http_image_urls_reachable is None


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_html_embeds_both_http_images_as_data_uri(
    public_http_image_urls_reachable,
    network_writer: FileWriter,
) -> None:
    md = (
        "# Отчёт\n\n"
        f"Первое фото:\n\n![]({IMG_HTTP_JPEG_1})\n\n"
        f"Второе фото:\n\n![]({IMG_HTTP_JPEG_2})\n\n"
        "Конец."
    )
    result = network_writer.build_bytes(md, "report.html", content_mode="markdown")
    overwrite_artifact("network_report.html", result.data)
    assert result.mime_type.startswith("text/html")
    html = result.data.decode("utf-8")
    assert "Отчёт" in html and "Конец." in html
    n = _count_data_uri_images(html)
    assert n >= 2, f"ожидались минимум 2 data:image в HTML, найдено {n}"


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_docx_contains_both_images_in_word_media(
    public_http_image_urls_reachable,
    network_writer: FileWriter,
) -> None:
    md = (
        "Введение\n\n"
        f"![]({IMG_HTTP_JPEG_1})\n\n"
        "Раздел 2\n\n"
        f"![]({IMG_HTTP_JPEG_2})\n\n"
        "Заключение."
    )
    result = network_writer.build_bytes(md, "doc.docx", content_mode="markdown")
    overwrite_artifact("network_doc.docx", result.data)
    assert "wordprocessingml" in result.mime_type
    zf = zipfile.ZipFile(BytesIO(result.data))
    media = [n for n in zf.namelist() if n.startswith("word/media/")]
    assert len(media) >= 2, f"ожидались минимум 2 файла в word/media/, получено {media}"
    doc_xml = zf.read("word/document.xml").decode("utf-8")
    assert "Введение" in doc_xml and "Заключение" in doc_xml


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_xlsx_table_text_and_both_http_images(
    public_http_image_urls_reachable,
    network_writer: FileWriter,
) -> None:
    md = (
        "Заголовок отчёта\n\n"
        "| Колонка | Значение |\n"
        "|---------|----------|\n"
        "| alpha   | 100      |\n"
        "| beta    | 200      |\n\n"
        f"Снимок A:\n\n![]({IMG_HTTP_JPEG_1})\n\n"
        f"Снимок B:\n\n![]({IMG_HTTP_JPEG_2})\n\n"
        "Итог: ок."
    )
    result = network_writer.build_bytes(md, "book.xlsx", content_mode="markdown")
    overwrite_artifact("network_book.xlsx", result.data)
    assert "spreadsheetml" in result.mime_type
    zf = zipfile.ZipFile(BytesIO(result.data))
    media = [n for n in zf.namelist() if n.startswith("xl/media/")]
    assert len(media) >= 2, f"ожидались минимум 2 в xl/media/, получено {media}"
    wb = load_workbook(BytesIO(result.data), read_only=True, data_only=True)
    ws = wb.active
    flat = " ".join(
        str(c.value)
        for row in ws.iter_rows()
        for c in row
        if c.value is not None
    )
    wb.close()
    assert "alpha" in flat and "Заголовок отчёта" in flat


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_pdf_larger_with_http_image_than_text_only(
    public_http_image_urls_reachable,
    network_writer: FileWriter,
) -> None:
    plain = network_writer.build_bytes(
        "# Только текст\n\nАбзац один. Абзац два.",
        "plain.pdf",
        content_mode="markdown",
    )
    with_img = network_writer.build_bytes(
        f"# С картинкой\n\n![]({IMG_HTTP_JPEG_1})\n\nПодпись.",
        "with_img.pdf",
        content_mode="markdown",
    )
    overwrite_artifact("network_plain.pdf", plain.data)
    overwrite_artifact("network_with_image.pdf", with_img.data)
    assert plain.data[:5] == b"%PDF-"
    assert with_img.data[:5] == b"%PDF-"
    assert len(with_img.data) > len(plain.data) + 800, (
        "PDF с реальной картинкой должен быть заметно больше текстового"
    )


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_md_file_preserves_markdown_and_both_image_urls(
    public_http_image_urls_reachable,
    network_writer: FileWriter,
) -> None:
    """Для .md конвертация не тянет байты по URL — в файле остаются исходные ссылки."""
    md = (
        f"# Заметка\n\n![]({IMG_HTTP_JPEG_1})\n\nещё\n\n![]({IMG_HTTP_JPEG_2})\n"
    )
    result = network_writer.build_bytes(md, "notes.md", content_mode="markdown")
    overwrite_artifact("network_notes.md", result.data)
    text = result.data.decode("utf-8")
    assert IMG_HTTP_JPEG_1 in text and IMG_HTTP_JPEG_2 in text
    assert "![](https://picsum.photos" in text


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_txt_file_preserves_markdown_and_both_image_urls(
    public_http_image_urls_reachable,
    network_writer: FileWriter,
) -> None:
    md = (
        f"Заголовок\n\n![]({IMG_HTTP_JPEG_1})\n\nТекст\n\n![]({IMG_HTTP_JPEG_2})\n"
    )
    result = network_writer.build_bytes(md, "export.txt", content_mode="markdown")
    overwrite_artifact("network_export.txt", result.data)
    text = result.data.decode("utf-8")
    assert IMG_HTTP_JPEG_1 in text and IMG_HTTP_JPEG_2 in text


@pytest.mark.integration
@pytest.mark.network
@pytest.mark.timeout(180, func_only=True)
def test_pdf_contains_both_http_images(
    public_http_image_urls_reachable,
    network_writer: FileWriter,
) -> None:
    """PDF с двумя внешними картинками — размер и валидный заголовок."""
    md = (
        f"![]({IMG_HTTP_JPEG_1})\n\n"
        f"![]({IMG_HTTP_JPEG_2})"
    )
    one = network_writer.build_bytes(
        f"![]({IMG_HTTP_JPEG_1})",
        "one.pdf",
        content_mode="markdown",
    )
    two = network_writer.build_bytes(md, "two.pdf", content_mode="markdown")
    overwrite_artifact("network_one_image.pdf", one.data)
    overwrite_artifact("network_two_images.pdf", two.data)
    assert two.data[:5] == b"%PDF-"
    assert len(two.data) > len(one.data) + 400, "две картинки дают больший PDF чем одна"
